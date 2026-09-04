#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Genere le contrat d'interface SAP -> IFS du module ARTICLES (inventory).

Meme structure que sql/supplier/contrat_interface_SAP_IFS_Supplier.xlsx :
  - un onglet Legende
  - un onglet 00_VALEURS_PAR_DEFAUT (contenu de public.etl_default_values)
  - un onglet par table cible, dans l'ordre d'execution du module

Le contenu n'est PAS saisi a la main : il est derive
  - des scripts alimenter_*.sql (colonne cible <-> expression <-> commentaire),
  - de la base (types reels via information_schema, valeurs par defaut,
    exemples de valeurs reellement chargees).

Usage :
    python generer_contrat_interface.py [chemin_sortie.xlsx]
"""

import os
import re
import sys

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Connexion (memes valeurs que compile.sh / export_procedures.sh, surchargeables)
# ---------------------------------------------------------------------------
DB = dict(
    host=os.getenv('DB_HOST', '10.190.100.58'),
    port=os.getenv('DB_PORT', '5432'),
    dbname=os.getenv('DB_NAME', 'sap_migration_db'),
    user=os.getenv('DB_USER', 'postgres'),
    password=os.getenv('DB_PASSWORD', 'trimet2025'),
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# Les 7 etapes du module, dans l'ordre d'execution (cf. compile.sh)
# ---------------------------------------------------------------------------
ETAPES = [
    ('01_IFS_ARTICLE_MAITRE', 'alimenter_ifs_article.sql', 'clean_data.ifs_article_maitre',
     "TABLE PILOTE. Perimetre = raw_data.export_article_qlikview (liste des articles a migrer), "
     "restreint aux articles ayant une fiche MARA active (mandt 700, non marques a supprimer). "
     "Agrege les donnees SAP par article : centres (MARC), commercial (MVKE), evaluation (MBEW), "
     "stocks (MARD), fournisseur principal (EINA/LFA1)."),
    ('02_PART_CATALOG', 'alimenter_part_catalog.sql', 'clean_data.part_catalog',
     "Catalogue des pieces IFS : table de BASE du module. Reprise 1:1 de ifs_article_maitre. "
     "Toutes les tables suivantes filtrent leur contenu sur la presence de l'article ici."),
    ('03_INVENTORY_PART', 'alimenter_inventory_part.sql', 'clean_data.inventory_part',
     "Article d'inventaire par site (contract). Un enregistrement par couple article x centre "
     "SAP 9200 (SJ) / 9000 (CS)."),
    ('04_INVENT_PART_PLAN', 'alimenter_inventory_part_planning.sql', 'clean_data.invent_part_plan',
     "Parametres de planification (lot, stock de securite, point de commande) par article x site. "
     "Filtre sur la presence dans inventory_part."),
    ('05_PURCHASE_PART', 'alimenter_purchase_part.sql', 'clean_data.purchase_part',
     "Article d'achat par site. Centres 9200 / 9000 / 2200 et approvisionnement externe "
     "(MARC.BESKZ dans F, X). Les articles de vente (raw_data.articles_vente_sap) en sont EXCLUS : "
     "ils relevent de SALES_PART."),
    ('06_PURCHASE_PART_SUPPLIER', 'alimenter_purchase_part_supplier.sql',
     'clean_data.purchase_part_supplier',
     "Lien article x fournisseur issu des fiches info-achat SAP (EINA/EINE). Le fournisseur est "
     "remappe sur son NOUVEAU numero IFS (600xxx) via supplier_info_general.supplier_legacy_sap_id."),
    ('07_SALES_PART', 'alimenter_sales_part.sql', 'clean_data.sales_part',
     "Article de vente. Perimetre = raw_data.articles_vente_sap, restreint aux types d'article "
     "FERT / HALB / DIEN / NLAG / HIBE / ERSA presents dans part_catalog."),
]

# ---------------------------------------------------------------------------
# Styles (identiques au contrat fournisseurs)
# ---------------------------------------------------------------------------
BLEU_ENTETE = PatternFill('solid', fgColor='1F4E78')
BLEU_SECTION = PatternFill('solid', fgColor='D9E1F2')
JAUNE_METIER = PatternFill('solid', fgColor='FFF2CC')
GRIS_INACTIF = PatternFill('solid', fgColor='F2F2F2')
F_ENTETE = Font(bold=True, size=10, color='FFFFFF')
F_SECTION = Font(bold=True, size=10)
F_CELL = Font(size=10)
BORD = Border(*[Side(style='thin', color='BFBFBF')] * 4)
AL_ENTETE = Alignment(wrap_text=True, vertical='center')
AL_CELL = Alignment(wrap_text=True, vertical='top')

ENTETES = ['Colonne cible', 'Type / Longueur', 'Champ / Table source', 'Systeme source',
           'Regle de transformation / Condition', 'Exemple de valeur',
           'Valide metier (O/N)', 'Remarques metier', 'Date validation']
LARGEURS = [30, 15, 34, 14, 62, 20, 15, 32, 13]

# ---------------------------------------------------------------------------
# Parsing des scripts SQL
# ---------------------------------------------------------------------------
QUOTES = "'" + '"'


def _scan(txt, kw, depart=0):
    """Position du 1er mot-cle kw a profondeur 0 (hors quotes et commentaires)."""
    depth, i, q, n = 0, depart, None, len(txt)
    while i < n:
        c = txt[i]
        if q:
            if c == q:
                q = None
        elif txt[i:i + 2] == '--':
            j = txt.find(chr(10), i)
            if j < 0:
                break
            i = j
        elif c in QUOTES:
            q = c
        elif c == '(':
            depth += 1
        elif c == ')':
            depth -= 1
        elif depth == 0 and txt[i:i + len(kw)].upper() == kw:
            avant = txt[i - 1] if i else ' '
            apres = txt[i + len(kw)] if i + len(kw) < n else ' '
            if not (avant.isalnum() or avant == '_') and not (apres.isalnum() or apres == '_'):
                return i
        i += 1
    return -1


def _decouper(txt):
    """Decoupe sur les virgules de niveau 0 (commentaires et quotes preserves)."""
    parts, depth, cur, i, q = [], 0, [], 0, None
    while i < len(txt):
        c = txt[i]
        if q:
            cur.append(c)
            if c == q:
                q = None
        elif txt[i:i + 2] == '--':
            j = txt.find(chr(10), i)
            j = len(txt) if j < 0 else j
            cur.append(txt[i:j])
            i = j
            continue
        elif c in QUOTES:
            q = c
            cur.append(c)
        elif c == '(':
            depth += 1
            cur.append(c)
        elif c == ')':
            depth -= 1
            cur.append(c)
        elif c == ',' and depth == 0:
            parts.append(''.join(cur))
            cur = []
        else:
            cur.append(c)
        i += 1
    if ''.join(cur).strip():
        parts.append(''.join(cur))
    return parts


def _sans_commentaire(s):
    return re.sub(r'--[^\n]*', '', s)


def parser_script(chemin):
    """Retourne la liste [(colonne_cible, expression, commentaire)] du INSERT ... SELECT."""
    src = open(chemin, encoding='utf-8').read()
    m = re.search(r'INSERT\s+INTO\s+(\S+)\s*\(', src, re.I)
    if not m:
        raise ValueError('INSERT introuvable dans %s' % chemin)

    ouvrante = m.end() - 1
    depth, i = 0, ouvrante
    while i < len(src):
        if src[i] == '(':
            depth += 1
        elif src[i] == ')':
            depth -= 1
            if depth == 0:
                break
        i += 1
    colonnes = [c.strip() for c in
                (_sans_commentaire(x) for x in _decouper(src[ouvrante + 1:i])) if c.strip()]

    reste = src[i + 1:]
    debut = _scan(reste, 'SELECT') + len('SELECT')
    md = re.match(r'\s+DISTINCT(\s+ON\s*\([^)]*\))?', reste[debut:], re.I)
    if md:
        debut += md.end()
    fin = _scan(reste, 'FROM', debut)
    corps, bloc_from = reste[debut:fin], reste[fin:]

    expressions = []
    for e in _decouper(corps):
        commentaires = [c.strip() for c in re.findall(r'--\s*(.+)', e)
                        if not re.match(r'^=+$', c.strip())]
        expressions.append((' '.join(_sans_commentaire(e).split()),
                            ' '.join(commentaires).strip()))

    if len(colonnes) != len(expressions):
        raise ValueError('%s : %d colonnes pour %d expressions (INSERT/SELECT desynchronises)'
                         % (os.path.basename(chemin), len(colonnes), len(expressions)))

    return ([(c, e, k) for c, (e, k) in zip(colonnes, expressions)],
            _alias_du_from(bloc_from), parser_ctes(src))


def _alias_du_from(bloc_from):
    """alias -> table, lu dans le FROM/JOIN."""
    alias = {}
    for tm in re.finditer(r'(?:FROM|JOIN)\s+([a-z_]+\.[a-z_0-9]+|[a-z_0-9]+)(?:\s+(?:AS\s+)?([a-z_0-9]+))?',
                          _sans_commentaire(bloc_from), re.I):
        table, al = tm.group(1), (tm.group(2) or '')
        if al.lower() in ('on', 'where', 'inner', 'left', 'right', 'join', 'order', 'group', 'and'):
            al = ''
        alias[(al or table.split('.')[-1]).lower()] = table
    return alias


def parser_ctes(src):
    """{nom_cte: {colonne_sortie: [tables.colonnes sources]}} pour resoudre les WITH."""
    ctes = {}
    for m in re.finditer(r'([a-z_0-9]+)\s+AS\s*\(', src, re.I):
        nom = m.group(1).lower()
        if nom in ('function', 'exists'):
            continue
        depth, i = 0, m.end() - 1
        while i < len(src):
            if src[i] == '(':
                depth += 1
            elif src[i] == ')':
                depth -= 1
                if depth == 0:
                    break
            i += 1
        corps_cte = src[m.end():i]
        pos_sel = _scan(corps_cte, 'SELECT')
        if pos_sel < 0:
            continue
        debut = pos_sel + len('SELECT')
        md = re.match(r'\s+DISTINCT(\s+ON\s*\([^)]*\))?', corps_cte[debut:], re.I)
        if md:
            debut += md.end()
        fin = _scan(corps_cte, 'FROM', debut)
        if fin < 0:
            continue
        alias_cte = _alias_du_from(corps_cte[fin:])
        cols = {}
        for e in _decouper(corps_cte[debut:fin]):
            code = ' '.join(_sans_commentaire(e).split())
            if not code:
                continue
            ma = re.search(r'\bAS\s+([a-z_0-9]+)\s*$', code, re.I)
            sortie = (ma.group(1) if ma else code.split('.')[-1]).lower()
            refs = []
            for a, col in re.findall(r'\b([a-z_0-9]+)\.([a-z_0-9]+)\b', code):
                t = alias_cte.get(a.lower())
                if t and '%s.%s' % (t, col) not in refs:
                    refs.append('%s.%s' % (t, col))
            if not refs:
                # agregat sans colonne source (count(*)...) : rattacher a la table du FROM
                principale = next(iter(alias_cte.values()), None)
                if principale:
                    refs = ['%s (agregat)' % principale]
            cols[sortie] = refs
        ctes[nom] = cols
    return ctes


# ---------------------------------------------------------------------------
# Derivation source / systeme / regle
# ---------------------------------------------------------------------------
def _resoudre(table, col, ctes, profondeur=0):
    """Remonte les CTE jusqu'aux tables reelles (raw_data / clean_data)."""
    nom = table.split('.')[-1].lower()
    if profondeur > 3 or nom not in ctes:
        return ['%s.%s' % (table, col)]
    refs = ctes[nom].get(col.lower())
    if not refs:
        return ['%s.%s' % (table, col)]
    sorties = []
    for r in refs:
        t, c = r.rsplit('.', 1)
        for s in _resoudre(t, c, ctes, profondeur + 1):
            if s not in sorties:
                sorties.append(s)
    return sorties


def champs_source(expr, alias, ctes):
    """Liste des 'schema.table.colonne' referencees par l'expression, CTE resolues."""
    trouves = []
    for a, col in re.findall(r'\b([a-z_0-9]+)\.([a-z_0-9]+)\b', expr, re.I):
        a = a.lower()
        if a in ('public', 'clean_data', 'raw_data'):
            continue
        table = alias.get(a, a)
        for ref in _resoudre(table, col, ctes):
            if ref not in trouves:
                trouves.append(ref)
    if 'get_default_value' in expr:
        for t, c in re.findall(r"get_default_value\(\s*'([^']+)'\s*,\s*'([^']+)'", expr):
            ref = 'public.etl_default_values (%s.%s)' % (t.split('.')[-1], c)
            if ref not in trouves:
                trouves.append(ref)
    if 'get_transcodification' in expr:
        for cat in re.findall(r"get_transcodification\(\s*'([^']+)'", expr):
            ref = 'public."TranscodificationTable" (categorie %s)' % cat
            if ref not in trouves:
                trouves.append(ref)
    return trouves


def systeme_source(expr, refs):
    sys_ = []
    if any(r.startswith('raw_data.export_article_qlikview') for r in refs):
        sys_.append('QlikView')
    if any(r.startswith('raw_data.') and not r.startswith('raw_data.export_article_qlikview')
           for r in refs):
        sys_.append('SAP')
    if any(r.startswith('clean_data.') for r in refs):
        sys_.append('IFS')
    if 'get_default_value' in expr or 'get_transcodification' in expr:
        sys_.append('Config')
    if not sys_:
        sys_.append('Technique')
    return ' / '.join(dict.fromkeys(sys_))


def regle(expr, commentaire):
    """Texte lisible : commentaire du script s'il existe, sinon l'expression."""
    expr_courte = expr if len(expr) <= 260 else expr[:257] + '...'
    com = re.sub(r'^[A-Z_0-9]+\s*:\s*', '', commentaire).strip()
    if com and len(com) > 12:
        if re.search(r'CASE|COALESCE|SUBSTRING|get_default_value|get_transcodification', expr, re.I):
            return '%s\nSQL : %s' % (com, expr_courte)
        return com
    return expr_courte


# ---------------------------------------------------------------------------
# Lecture base
# ---------------------------------------------------------------------------
def lire_base(cur, tables):
    types, exemples, defauts = {}, {}, []

    cur.execute("""
        SELECT table_name, column_name,
               CASE data_type
                 WHEN 'character varying' THEN 'VARCHAR' || COALESCE('('||character_maximum_length||')','')
                 WHEN 'character' THEN 'CHAR(' || character_maximum_length || ')'
                 WHEN 'numeric' THEN 'NUMERIC' || COALESCE('('||numeric_precision||','||numeric_scale||')','')
                 WHEN 'timestamp without time zone' THEN 'TIMESTAMP'
                 WHEN 'double precision' THEN 'FLOAT8'
                 ELSE upper(data_type) END
        FROM information_schema.columns
        WHERE table_schema = 'clean_data' AND table_name = ANY(%s)
    """, ([t.split('.')[1] for t in tables],))
    for t, c, ty in cur.fetchall():
        types['clean_data.%s.%s' % (t, c)] = ty

    # 1 exemple non NULL par colonne, sur un echantillon de lignes reellement chargees
    for t in tables:
        try:
            cur.execute('SELECT * FROM %s LIMIT 300' % t)
        except psycopg2.Error:
            cur.connection.rollback()
            continue
        noms = [d[0] for d in cur.description]
        vus = {}
        for ligne in cur.fetchall():
            for nom, val in zip(noms, ligne):
                if nom not in vus and val is not None and str(val).strip() != '':
                    vus[nom] = str(val)[:60]
            if len(vus) == len(noms):
                break
        for nom, val in vus.items():
            exemples['%s.%s' % (t, nom)] = val

    cur.execute("""
        SELECT table_cible, colonne, COALESCE(variante,'STANDARD'), type_valeur,
               COALESCE(valeur, '(NULL)'), is_active, module, description
        FROM public.etl_default_values
        WHERE table_cible = ANY(%s)
        ORDER BY table_cible, colonne, variante
    """, (tables,))
    defauts = cur.fetchall()
    return types, exemples, defauts


# ---------------------------------------------------------------------------
# Ecriture du classeur
# ---------------------------------------------------------------------------
def ligne_section(ws, texte, nb_col=9):
    r = ws.max_row + 1
    for c in range(1, nb_col + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = BLEU_SECTION
        if c == 1:
            cell.value = texte
            cell.font = F_SECTION
            cell.alignment = AL_ENTETE
    ws.row_dimensions[r].height = 18
    return r


def entete(ws, entetes, largeurs):
    ws.append(entetes)
    for i, (txt, w) in enumerate(zip(entetes, largeurs), 1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font, c.alignment, c.border = BLEU_ENTETE, F_ENTETE, AL_ENTETE, BORD
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'


def appels_defaut(lignes):
    """{(table_cible, colonne, variante)} reellement appeles via get_default_value."""
    cles = set()
    for _, expr, _ in lignes:
        for m in re.finditer(r"get_default_value\(\s*'([^']+)'\s*,\s*'([^']+)'"
                             r"(?:\s*,\s*(?:'[^']*'|[^,)]+))?(?:\s*,\s*'([^']+)')?", expr):
            cles.add((m.group(1), m.group(2), (m.group(3) or 'STANDARD')))
    return cles


def onglet_table(wb, onglet, script, table, resume, types, exemples, appels):
    lignes, alias, ctes = parser_script(os.path.join(BASE_DIR, script))
    appels |= appels_defaut(lignes)
    ws = wb.create_sheet(onglet)
    entete(ws, ENTETES, LARGEURS)
    ligne_section(ws, 'Script %s  —  %s' % (script, resume))

    for col, expr, com in lignes:
        refs = champs_source(expr, alias, ctes)
        r = ws.max_row + 1
        valeurs = [col,
                   types.get('%s.%s' % (table, col), ''),
                   ' ; '.join(refs) if refs else '—',
                   systeme_source(expr, refs),
                   regle(expr, com),
                   exemples.get('%s.%s' % (table, col), '')]
        for i, v in enumerate(valeurs, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.font, c.alignment, c.border = F_CELL, AL_CELL, BORD
        for i in (7, 8, 9):
            c = ws.cell(row=r, column=i)
            c.fill, c.font, c.alignment, c.border = JAUNE_METIER, F_CELL, AL_CELL, BORD
        ws.row_dimensions[r].height = 34
    return len(lignes)


def onglet_valeurs_defaut(wb, defauts, appels):
    ws = wb.create_sheet('00_VALEURS_PAR_DEFAUT')
    ent = ['Onglet', 'Table cible (etl_default_values.table_cible)', 'Colonne', 'Variante',
           'Type valeur', 'Valeur actuelle', 'Actif', 'Module', 'Description', 'Remarque metier']
    entete(ws, ent, [26, 34, 26, 16, 12, 22, 8, 12, 40, 30])

    # Ne garder que les parametres reellement appeles par les 7 scripts du module.
    # Les memes tables cibles portent aussi des variantes du module articlePhl,
    # qui ne relevent pas de ce contrat.
    par_table = {}
    for d in defauts:
        if (d[0], d[1], d[2]) in appels:
            par_table.setdefault(d[0], []).append(d)
    onglet_de = {t: o for o, _, t, _ in ETAPES}

    for onglet, _, table, _ in ETAPES:
        if table not in par_table:
            continue
        ligne_section(ws, '%s  —  %s' % (onglet, table), 10)
        for tbl, col, var, typ, val, actif, module, desc in par_table[table]:
            r = ws.max_row + 1
            for i, v in enumerate([onglet_de.get(tbl, ''), tbl, col, var, typ, val,
                                   'Oui' if actif else 'Non', module, desc or ''], 1):
                c = ws.cell(row=r, column=i, value=v)
                c.font, c.alignment, c.border = F_CELL, AL_CELL, BORD
                if not actif:
                    c.fill = GRIS_INACTIF
            c = ws.cell(row=r, column=10)
            c.fill, c.font, c.alignment, c.border = JAUNE_METIER, F_CELL, AL_CELL, BORD
    return sum(len(v) for v in par_table.values())


LEGENDE = [
    ("Contrat d'interface — Module Articles (SAP ECC 6.0 → IFS)", 14, True),
    (None, None, None),
    ("Objet : documenter, pour chacune des 7 etapes du chargement des articles (module ETL "
     "\"Donnees de base des Articles\", backend/etl_modules/etl_inventory_part.py), l'origine de "
     "chaque colonne cible et la regle de transformation appliquee, afin que le metier valide le "
     "mapping. Chaque onglet correspond a une table clean_data et au script qui l'alimente, dans "
     "l'ordre d'execution.", 10, False),
    (None, None, None),
    ("Structure d'un onglet table", 12, True),
    (None, None, None),
    ('• Colonne cible : Nom de la colonne dans la table clean_data (cible finale).', 10, False),
    (None, None, None),
    ('• Type / Longueur : Type et longueur reels, lus directement dans la base clean_data.', 10, False),
    (None, None, None),
    ('• Champ / Table source : Table(s)/champ(s) source(s) (raw_data SAP, clean_data amont, ou constante).', 10, False),
    (None, None, None),
    ("• Systeme source : SAP, QlikView (fichier de perimetre des articles), IFS (table clean_data "
     "amont), Config (etl_default_values / TranscodificationTable), ou Technique.", 10, False),
    (None, None, None),
    ("• Regle de transformation / Condition : COALESCE, CASE, concatenation, valeur par defaut... "
     "Colonne cle a relire cote metier.", 10, False),
    (None, None, None),
    ("• Exemple de valeur : valeur reellement presente en base apres le dernier chargement.", 10, False),
    (None, None, None),
    ('• Valide metier (O/N) / Remarques metier / Date validation : Cellules en jaune, a remplir par le metier.', 10, False),
    (None, None, None),
    (None, None, None),
    ("Les 7 etapes (dans l'ordre d'execution)", 12, True),
    (None, None, None),
]

ATTENTION = [
    ("Points d'attention transverses reperes (a faire trancher par le metier)", 12, True),
    (None, None, None),
    ("⚠ Perimetre — La liste des articles a migrer est pilotee par raw_data.export_article_qlikview "
     "(18 652 articles). Cette table n'est alimentee par AUCUN import applicatif : elle est chargee "
     "a la main. Si elle est vide au moment d'un chargement, les 7 tables du module se vident en "
     "cascade sans erreur.", 10, False),
    ("⚠ 01 — codification_id et new_transco ne sont plus alimentees : la table "
     "clean_data.mapping_codification_articles a ete retiree du mapping sur demande metier, et "
     "aucune source de remplacement n'est branchee. Les deux colonnes sortent a NULL.", 10, False),
    ("⚠ 01 — 89 articles du perimetre n'ont pas de libelle SAP (MAKT) en F/E/D : leur designation "
     "provient de export_article_qlikview. Les tables 03 et 05 exigeant MAKT en francais, ces "
     "articles n'apparaitront pas en article d'inventaire ni en article d'achat.", 10, False),
    ("⚠ 05 — Les articles de vente (raw_data.articles_vente_sap) sont exclus de PURCHASE_PART "
     "(264 articles) : achat et vente sont disjoints. A confirmer si certains articles doivent etre "
     "a la fois achetes et vendus.", 10, False),
    ("⚠ 06 — primary_vendor_db vaut 'Y' sur TOUTES les lignes, alors que 6 812 couples "
     "(contract, part_no) ont plusieurs fournisseurs (jusqu'a 12). IFS n'admet qu'un seul "
     "fournisseur principal par article et par site : le chargement rejettera ou ecrasera. "
     "Un seul 'Y' par couple doit etre designe (par exemple la fiche info-achat la plus recente).", 10, False),
    ("⚠ 06 — 3 744 liens article/fournisseur SAP sont perdus car le fournisseur (250 LIFNR "
     "distincts) n'existe pas dans clean_data.supplier_info_general. Le INNER JOIN est volontaire "
     "(un lien vers un fournisseur non migre serait rejete au chargement IFS), mais cela represente "
     "11 % des liens : a valider cote metier.", 10, False),
    ("⚠ Couples <colonne> / <colonne>_db : par convention du depot, seule la colonne _db est "
     "renseignee. Les colonnes sans suffixe restent NULL.", 10, False),
    (None, None, None),
    ('Onglet 00_VALEURS_PAR_DEFAUT', 12, True),
    (None, None, None),
]


def onglet_legende(wb, nb_defauts, compte_lignes):
    ws = wb.create_sheet('Legende', 0)
    ws.column_dimensions['A'].width = 100

    def ecrire(txt, taille, gras):
        r = ws.max_row + 1 if ws.max_row > 1 or ws['A1'].value else 1
        c = ws.cell(row=r, column=1, value=txt)
        if txt is not None:
            c.font = Font(bold=gras, size=taille)
            c.alignment = Alignment(wrap_text=not gras)

    for txt, taille, gras in LEGENDE:
        ecrire(txt, taille, gras)
    for i, (onglet, script, table, resume) in enumerate(ETAPES, 1):
        ecrire('%02d. %s — %s (%s colonnes). %s' % (i, onglet.split('_', 1)[1], table,
                                                    compte_lignes.get(onglet, '?'), resume),
               10, False)
    ecrire(None, None, None)
    for txt, taille, gras in ATTENTION:
        ecrire(txt, taille, gras)
    ecrire("Liste exhaustive (%d lignes) des valeurs enregistrees dans public.etl_default_values "
           "pour les 7 tables du module articles — celles que les scripts appellent via "
           "public.get_default_value(). Regroupee par onglet/table puis par colonne. Les lignes "
           "grisees correspondent a des variantes inactives (is_active = FALSE). C'est ici que la "
           "VALEUR reelle de chaque parametre est visible et modifiable "
           "(Configuration > Valeurs par defaut dans l'outil)." % nb_defauts, 10, False)


def main():
    sortie = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        BASE_DIR, 'contrat_interface_SAP_IFS_Articles.xlsx')

    tables = [t for _, _, t, _ in ETAPES]
    conn = psycopg2.connect(**DB)
    try:
        with conn.cursor() as cur:
            types, exemples, defauts = lire_base(cur, tables)
    finally:
        conn.close()

    wb = Workbook()
    wb.remove(wb.active)
    compte, appels = {}, set()
    for onglet, script, table, resume in ETAPES:
        compte[onglet] = onglet_table(wb, onglet, script, table, resume, types, exemples, appels)
        print('  %-28s %3d colonnes' % (onglet, compte[onglet]))
    nb_def = onglet_valeurs_defaut(wb, defauts, appels)
    print('  %-28s %3d lignes' % ('00_VALEURS_PAR_DEFAUT', nb_def))
    onglet_legende(wb, nb_def, compte)

    # ordre des onglets : Legende, 00, puis 01..07
    wb._sheets = ([wb['Legende'], wb['00_VALEURS_PAR_DEFAUT']] +
                  [wb[o] for o, _, _, _ in ETAPES])
    wb.save(sortie)
    print('\nOK : %s (%d colonnes documentees)' % (sortie, sum(compte.values())))


if __name__ == '__main__':
    main()
