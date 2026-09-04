# -*- coding: utf-8 -*-
"""
Lecture / ecriture du classeur "contrat d'interface SAP -> IFS".

Ce module est volontairement SANS dependance a la base : il ne manipule que
des dictionnaires. Il est utilise par
  - scripts/seed_interface_contracts.py  : reprise initiale du classeur v3
    (sql/supplier/contrat_interface_SAP_IFS_Supplier.xlsx) vers les tables
    public.interface_contract_* ;
  - backend/api/interface_contracts.py   : export a la demande (regeneration
    depuis l'etat reel de la base) et reimport d'un classeur rempli hors ligne
    par un relecteur sans compte applicatif.

Format du classeur (identique a celui produit par
sql/inventory/generer_contrat_interface.py) :
  - 1 onglet "Legende"
  - 1 onglet "00_VALEURS_PAR_DEFAUT" (contenu de public.etl_default_values)
  - 1 onglet par table cible, 9 colonnes, dont 3 colonnes "metier" jaunes
    (Valide O/N, Remarques, Date validation) qui sont la partie relue.

Encodage des types de ligne dans un onglet table (c'est le style qui porte
l'information, pas le contenu) :
  - fond bleu clair D9E1F2 + gras/italique, colonne A seule -> titre de
    section. Le TOUT PREMIER de l'onglet est la description de la table
    ("Script 02_alimenter_....sql - ...") et non une section.
  - italique sans fond, colonne A seule                     -> NOTE
  - ligne dont la source commence par "Config"              -> CONFIG_SUMMARY
  - le reste                                                -> COLUMN
"""

import re
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Styles (identiques au classeur v3 : la reprise et l'export doivent produire
# exactement le meme rendu, sinon le parsing du reimport ne retrouve plus ses
# petits)
# ---------------------------------------------------------------------------
BLEU_ENTETE = PatternFill('solid', fgColor='1F4E78')
BLEU_SECTION = PatternFill('solid', fgColor='D9E1F2')
JAUNE_METIER = PatternFill('solid', fgColor='FFF2CC')
VERT_VALIDE = PatternFill('solid', fgColor='E2EFDA')
ROUGE_CORRIGER = PatternFill('solid', fgColor='FCE4E4')
ORANGE_OBSOLETE = PatternFill('solid', fgColor='FFF2CC')

F_ENTETE = Font(bold=True, size=10, color='FFFFFF')
F_SECTION = Font(bold=True, italic=True, size=10)
F_NOTE = Font(italic=True, size=10)
F_CELL = Font(size=10)
BORD = Border(*[Side(style='thin', color='BFBFBF')] * 4)
AL_ENTETE = Alignment(wrap_text=True, vertical='center')
AL_CELL = Alignment(wrap_text=True, vertical='top')

RGB_SECTION = '00D9E1F2'

ENTETES = ['Colonne cible', 'Type / Longueur', 'Champ / Table source', 'Système source',
           'Règle de transformation / Condition', 'Exemple de valeur',
           'Validé métier (O/N)', 'Remarques métier', 'Date validation']
LARGEURS = [34, 15, 34, 16, 62, 20, 15, 34, 13]

ENTETES_CONFIG = ['Onglet', 'Table cible (etl_default_values.table_cible)', 'Colonne', 'Variante',
                  'Type valeur', 'Valeur actuelle', 'Actif', 'Module', 'Description',
                  'Remarque métier']
LARGEURS_CONFIG = [30, 40, 30, 18, 13, 24, 8, 14, 60, 34]

ONGLET_LEGENDE = 'Légende'
ONGLET_CONFIG = '00_VALEURS_PAR_DEFAUT'

# Correspondance "Validé métier (O/N)" <-> statut en base
_OUI = {'O', 'OUI', 'Y', 'YES', 'X', 'VALIDE', 'VALIDÉ'}
_NON = {'N', 'NON', 'KO', 'A CORRIGER', 'À CORRIGER', 'A_CORRIGER'}
_NA = {'NA', 'N/A', 'SO', 'SANS OBJET', 'NON_APPLICABLE', 'NON APPLICABLE'}

STATUT_VERS_EXCEL = {
    'VALIDE': 'O',
    'A_CORRIGER': 'N',
    'NON_APPLICABLE': 'N/A',
    'A_VALIDER': '',
}


def statut_depuis_excel(valeur):
    """'O'/'N'/'N/A'/vide -> statut en base, ou None si la case est vide."""
    if valeur is None:
        return None
    texte = str(valeur).strip().upper()
    if not texte:
        return None
    if texte in _OUI:
        return 'VALIDE'
    if texte in _NON:
        return 'A_CORRIGER'
    if texte in _NA:
        return 'NON_APPLICABLE'
    return None


# ---------------------------------------------------------------------------
# LECTURE
# ---------------------------------------------------------------------------
def _texte(cell):
    if cell is None or cell.value is None:
        return None
    texte = str(cell.value).strip()
    return texte or None


def _est_section(cell):
    fill = cell.fill
    rgb = getattr(getattr(fill, 'fgColor', None), 'rgb', None)
    return rgb == RGB_SECTION


def _mapping_onglets(wb):
    """Onglet -> nom qualifie de la table cible, lu sur les lignes de
    regroupement de 00_VALEURS_PAR_DEFAUT ('01_X  —  clean_data.x')."""
    mapping = {}
    if ONGLET_CONFIG not in wb.sheetnames:
        return mapping
    for row in wb[ONGLET_CONFIG].iter_rows(min_row=2, values_only=True):
        if row and row[0] and not row[1]:
            parties = re.split(r'\s+[—-]{1,2}\s+', str(row[0]), maxsplit=1)
            if len(parties) == 2:
                mapping[parties[0].strip()] = parties[1].strip()
    return mapping


def _libelle_onglet(nom):
    """'02_SUPPLIER_INFO_GENERAL' -> ('02 - SUPPLIER INFO GENERAL', 2)."""
    m = re.match(r'^(\d+)[_\-\s]+(.*)$', nom)
    if not m:
        return nom, 0
    return '%s - %s' % (m.group(1), m.group(2).replace('_', ' ')), int(m.group(1))


def _variante_depuis_libelle(libelle):
    """Extrait la ou les variantes etl_default_values citees dans le libelle
    d'une ligne CONFIG_SUMMARY : '... (variante TVA_UE/SIREN/SIRET)'.
    Retourne 'STANDARD' quand le libelle n'en mentionne aucune."""
    if not libelle:
        return 'STANDARD'
    m = re.search(r'variantes?\s+([A-Z0-9_/]+)', libelle, re.I)
    if m:
        return m.group(1).upper()
    return 'STANDARD'


def parse_contract_workbook(source, module='supplier'):
    """Lit un classeur de contrat et retourne la liste des tables, chacune
    portant ses lignes dans l'ordre du classeur.

    `source` : chemin ou objet fichier binaire.
    """
    wb = load_workbook(source, data_only=True)
    mapping = _mapping_onglets(wb)
    tables = []

    for nom in wb.sheetnames:
        if nom in (ONGLET_LEGENDE, ONGLET_CONFIG):
            continue
        ws = wb[nom]
        libelle, ordre = _libelle_onglet(nom)
        qualifie = mapping.get(nom)
        if qualifie and '.' in qualifie:
            schema_cible, table_cible = qualifie.split('.', 1)
        else:
            schema_cible = 'clean_data'
            table_cible = re.sub(r'^\d+[_\-]', '', nom).lower()

        table = {
            'sheet': nom,
            'module': module,
            'ordre': ordre,
            'libelle': libelle,
            'schema_cible': schema_cible,
            'table_cible': table_cible,
            'description': None,
            'source_procedure': None,
            'rows': [],
        }

        section = None
        sort_order = 0
        for cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
            premiere = _texte(cells[0])
            if premiere is None:
                continue
            valeurs = [_texte(c) for c in cells]
            renseignees = sum(1 for v in valeurs if v)

            # 1. bandeau bleu : description de la table (le premier) ou section
            if _est_section(cells[0]) and renseignees <= 1:
                if table['description'] is None:
                    table['description'] = premiere
                    m = re.search(r'Script\s+(\S+\.sql)', premiere)
                    if m:
                        table['source_procedure'] = m.group(1)
                else:
                    section = premiere
                continue

            sort_order += 10

            # 2. note libre (italique, colonne A seule)
            if cells[0].font and cells[0].font.italic and renseignees <= 1:
                table['rows'].append({
                    'section': section,
                    'target_column': '__NOTE_%d' % sort_order,
                    'row_type': 'NOTE',
                    'transformation_rule': premiere,
                    'sort_order': sort_order,
                    'default_value_variante': 'STANDARD',
                })
                continue

            source_expression = valeurs[2] if len(valeurs) > 2 else None
            systeme_source = valeurs[3] if len(valeurs) > 3 else None
            est_config = bool(
                (source_expression and source_expression.lower().startswith('config'))
                or (systeme_source and systeme_source.strip().lower() == 'config')
            )

            ligne = {
                'section': section,
                'target_column': premiere,
                'row_type': 'CONFIG_SUMMARY' if est_config else 'COLUMN',
                'systeme_source': systeme_source,
                'source_expression': source_expression,
                'transformation_rule': valeurs[4] if len(valeurs) > 4 else None,
                'exemple_valeur': valeurs[5] if len(valeurs) > 5 else None,
                'sort_order': sort_order,
                'default_value_variante': (
                    _variante_depuis_libelle(premiere) if est_config else 'STANDARD'
                ),
            }
            # colonnes metier : renseignees seulement si un relecteur a rempli
            # le classeur hors ligne (reimport)
            statut = statut_depuis_excel(valeurs[6] if len(valeurs) > 6 else None)
            remarque = valeurs[7] if len(valeurs) > 7 else None
            if statut or remarque:
                ligne['validation'] = {'statut': statut, 'remarque_metier': remarque}
            table['rows'].append(ligne)

        tables.append(table)

    return tables


# ---------------------------------------------------------------------------
# ECRITURE
# ---------------------------------------------------------------------------
LEGENDE = [
    "Contrat d'interface — Migration SAP ECC 6.0 -> IFS",
    '',
    "Objet : documenter, pour chaque table cible IFS, l'origine SAP de chaque",
    "colonne, la regle de transformation appliquee, et recueillir la validation",
    'metier.',
    '',
    "Ce classeur est GENERE a la demande depuis Migration Factory (ecran",
    "« Contrats d'interface »). La source de verite est l'application, pas ce",
    'fichier : une regeneration reflete toujours le dernier etat connu.',
    '',
    "Structure d'un onglet table",
    '',
    '• Colonne cible : colonne de la table IFS alimentee. Plusieurs colonnes',
    "  partageant la meme regle sont regroupees sur une seule ligne (« a / b »).",
    '• Type / Longueur : type reel de la colonne, lu dans la base.',
    '• Champ / Table source : table(s) et champ(s) SAP ou IFS a l origine.',
    '• Systeme source : SAP, IFS, Config (valeur par defaut parametrable) ou',
    '  Technique (constante posee par le chargement).',
    '• Regle de transformation / Condition : la logique appliquee.',
    "• Exemple de valeur : une valeur reelle, a titre d'illustration.",
    '',
    'Colonnes a remplir par le metier (fond jaune)',
    '',
    '• Valide metier (O/N) : O = valide, N = a corriger, N/A = sans objet.',
    '• Remarques metier : ce qui doit etre corrige, ou la precision utile.',
    '• Date validation : renseignee automatiquement a la reimportation.',
    '',
    'Un classeur rempli hors ligne peut etre reimporte dans Migration Factory :',
    "les statuts et remarques sont repris ligne a ligne, l'historique conserve",
    'le nom du relecteur.',
]


def _ecrire_legende(wb):
    ws = wb.create_sheet(ONGLET_LEGENDE)
    ws.column_dimensions['A'].width = 110
    for i, texte in enumerate(LEGENDE, 1):
        cell = ws.cell(row=i, column=1, value=texte)
        cell.font = Font(bold=(i == 1), size=11 if i == 1 else 10)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    return ws


def _ecrire_entete(ws, entetes, largeurs):
    for i, (titre, largeur) in enumerate(zip(entetes, largeurs), 1):
        cell = ws.cell(row=1, column=i, value=titre)
        cell.fill = BLEU_ENTETE
        cell.font = F_ENTETE
        cell.alignment = AL_ENTETE
        cell.border = BORD
        ws.column_dimensions[get_column_letter(i)].width = largeur
    ws.freeze_panes = 'A2'


def _bandeau(ws, ligne, texte, nb_colonnes):
    for col in range(1, nb_colonnes + 1):
        cell = ws.cell(row=ligne, column=col, value=texte if col == 1 else None)
        cell.fill = BLEU_SECTION
        cell.font = F_SECTION
        cell.alignment = AL_CELL
        cell.border = BORD


def _format_date(valeur):
    if valeur is None:
        return None
    if isinstance(valeur, (datetime, date)):
        return valeur.strftime('%Y-%m-%d')
    return str(valeur)[:10]


def build_contract_workbook(tables, default_values=None, types_colonnes=None):
    """Construit le classeur a partir de l'etat de la base.

    tables          : liste de dicts {libelle, table_cible, schema_cible,
                      description, sheet, rows: [...]} — rows au format de la
                      vue v_interface_contract.
    default_values  : liste de dicts etl_default_values (onglet 00), optionnel.
    types_colonnes  : {(schema, table, colonne): 'varchar(100)'} pour remplir
                      la colonne "Type / Longueur", optionnel.
    """
    types_colonnes = types_colonnes or {}
    wb = Workbook()
    wb.remove(wb.active)
    _ecrire_legende(wb)

    if default_values is not None:
        _ecrire_onglet_config(wb, tables, default_values)

    for table in tables:
        ws = wb.create_sheet(_nom_onglet(table))
        _ecrire_entete(ws, ENTETES, LARGEURS)
        ligne = 2
        if table.get('description'):
            _bandeau(ws, ligne, table['description'], len(ENTETES))
            ligne += 1
        section_courante = None

        for row in table['rows']:
            section = row.get('section')
            if section and section != section_courante:
                _bandeau(ws, ligne, section, len(ENTETES))
                ligne += 1
                section_courante = section

            if row.get('row_type') == 'NOTE':
                cell = ws.cell(row=ligne, column=1, value=row.get('transformation_rule'))
                cell.font = F_NOTE
                cell.alignment = AL_CELL
                for col in range(1, len(ENTETES) + 1):
                    ws.cell(row=ligne, column=col).border = BORD
                ligne += 1
                continue

            types = [
                types_colonnes.get(
                    (table.get('schema_cible'), table.get('table_cible'), nom.strip())
                )
                for nom in str(row.get('target_column') or '').split('/')
            ]
            type_longueur = ' / '.join(t for t in types if t) or None

            valeurs = [
                row.get('target_column'),
                type_longueur,
                row.get('source_expression'),
                row.get('systeme_source'),
                row.get('transformation_rule'),
                row.get('exemple_valeur'),
                STATUT_VERS_EXCEL.get(row.get('statut') or 'A_VALIDER', ''),
                row.get('remarque_metier'),
                _format_date(row.get('validated_at')),
            ]
            for col, valeur in enumerate(valeurs, 1):
                cell = ws.cell(row=ligne, column=col, value=valeur)
                cell.font = F_CELL
                cell.alignment = AL_CELL
                cell.border = BORD
                if col >= 7:
                    cell.fill = JAUNE_METIER
            # relecture d'un coup d'oeil : vert = valide, rouge = a corriger,
            # orange = valide mais la regle a change depuis
            if row.get('validation_obsolete'):
                ws.cell(row=ligne, column=7).fill = ORANGE_OBSOLETE
            elif row.get('statut') == 'VALIDE':
                ws.cell(row=ligne, column=7).fill = VERT_VALIDE
            elif row.get('statut') == 'A_CORRIGER':
                ws.cell(row=ligne, column=7).fill = ROUGE_CORRIGER
            ligne += 1

    return wb


def _nom_onglet(table):
    """Nom d'onglet Excel : 31 caracteres maxi, pas de caractere interdit."""
    brut = table.get('sheet') or '%02d_%s' % (table.get('ordre') or 0,
                                              (table.get('table_cible') or '').upper())
    return re.sub(r'[\\/*?:\[\]]', '_', brut)[:31]


def _ecrire_onglet_config(wb, tables, default_values):
    """Onglet 00 : toutes les valeurs par defaut parametrables, regroupees par
    onglet/table, dans le meme format que le classeur v3."""
    ws = wb.create_sheet(ONGLET_CONFIG)
    _ecrire_entete(ws, ENTETES_CONFIG, LARGEURS_CONFIG)
    par_table = {}
    for dv in default_values:
        par_table.setdefault(dv['table_cible'], []).append(dv)

    ligne = 2
    for table in tables:
        qualifie = '%s.%s' % (table.get('schema_cible'), table.get('table_cible'))
        lignes = par_table.get(qualifie)
        if not lignes:
            continue
        _bandeau(ws, ligne, '%s  —  %s' % (_nom_onglet(table), qualifie), len(ENTETES_CONFIG))
        ligne += 1
        for dv in sorted(lignes, key=lambda d: (d.get('colonne') or '', d.get('variante') or '')):
            valeurs = [
                _nom_onglet(table), qualifie, dv.get('colonne'), dv.get('variante'),
                dv.get('type_valeur'),
                '(NULL)' if dv.get('type_valeur') == 'NULL' else dv.get('valeur'),
                'Oui' if dv.get('is_active') else 'Non',
                dv.get('module'), dv.get('description'), None,
            ]
            for col, valeur in enumerate(valeurs, 1):
                cell = ws.cell(row=ligne, column=col, value=valeur)
                cell.font = F_CELL
                cell.alignment = AL_CELL
                cell.border = BORD
                if col == len(ENTETES_CONFIG):
                    cell.fill = JAUNE_METIER
                elif not dv.get('is_active'):
                    cell.fill = PatternFill('solid', fgColor='F2F2F2')
            ligne += 1
    return ws
